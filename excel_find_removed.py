# take two excel sheets, an original sheet and a copy of that one except with deleted entries.
# this program finds the deleted entries and adds them to a new sheet.

import openpyxl

orig_wkbk_path = "C:\\Users\\arkad\\Desktop\\Corbett\\Customers__Service_Titan.xlsx"
mod_wkbk_path = "C:\\Users\\arkad\Desktop\Corbett\Service_Titan_Nondeleted_Customers.xlsx"
columns = 13
orig_rows = 1009
mod_rows = 555
is_title = True

orig_wkbk = openpyxl.load_workbook(orig_wkbk_path)
mod_wkbk = openpyxl.load_workbook(mod_wkbk_path)

orig_sheet = orig_wkbk.active
mod_sheet = mod_wkbk.active

orig_content = []
mod_content = []
deleted_content = []
for aa in range(2, orig_rows + 1):
    content = []
    for bb in range(1, columns + 1):
        content.append(orig_sheet.cell(row=aa, column=bb).value)
    orig_content.append(content)

for aa in range(2, mod_rows + 1):
    content = []
    for bb in range(1, columns + 1):
        content.append(mod_sheet.cell(row=aa, column=bb).value)
    mod_content.append(content)

for cc in orig_content:
    if cc not in mod_content:
        deleted_content.append(cc)

workbook = openpyxl.Workbook()
sheet = workbook.active

if is_title:
    title = []
    for aa in range(1, columns + 1):
        title.append(orig_sheet.cell(row=1, column=aa).value)
    for aa in range(columns):
        sheet.cell(row=1, column=aa+1).value = title[aa]

for aa in range(2, len(deleted_content) + 1):
    for bb in range(1, columns + 1):
        sheet.cell(row=aa, column=bb).value = deleted_content[aa - 1][bb - 1]

workbook.save(filename="sample.xlsx")
